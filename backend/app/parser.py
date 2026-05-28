"""Parser do Gabarito_questoes.xlsx (agregado por questão, nível geral).

Aba esperada: "Gabarito". Cabeçalhos esperados (case-insensitive):
  Número da questão | Disciplina | Acertos Escola | Gabarito |
  A | B | C | D | E | Item Mais Marcado | Segundo Item Mais Marcado | Terceiro Item Mais Marcado
"""
import io
from typing import Optional

from openpyxl import load_workbook

# header normalizado -> chave interna
HEADER_MAP = {
    "número da questão": "numero",
    "numero da questao": "numero",
    "disciplina": "disciplina",
    "acertos escola": "acerto",
    "gabarito": "gabarito",
    "a": "A", "b": "B", "c": "C", "d": "D", "e": "E",
    "item mais marcado": "item_1",
    "segundo item mais marcado": "item_2",
    "terceiro item mais marcado": "item_3",
}

REQUIRED = {"numero", "disciplina", "acerto", "gabarito", "A", "B", "C", "D", "E"}
ALTERNATIVAS = {"A", "B", "C", "D", "E"}


def _norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def _alt(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip().upper()
    return s if s in ALTERNATIVAS else None


def _int(v) -> int:
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


def parse_gabarito(content: bytes):
    """Retorna (questoes, erros). 'questoes' é uma lista de dicts prontos para o banco.

    Levanta ValueError em problemas estruturais (aba/colunas ausentes).
    Erros por-linha são acumulados em 'erros' sem abortar o import.
    """
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb["Gabarito"] if "Gabarito" in wb.sheetnames else wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Planilha vazia.")

    header = [_norm(c) for c in rows[0]]
    col = {HEADER_MAP[h]: i for i, h in enumerate(header) if h in HEADER_MAP}
    missing = REQUIRED - col.keys()
    if missing:
        # mapeia de volta para os nomes amigáveis
        nomes = {v: k for k, v in HEADER_MAP.items()}
        raise ValueError("Colunas ausentes: " + ", ".join(sorted(nomes.get(m, m) for m in missing)))

    questoes, erros = [], []
    for ridx, row in enumerate(rows[1:], start=2):
        if not row or all(v is None for v in row):
            continue  # linha totalmente vazia
        numero_raw = row[col["numero"]] if col["numero"] < len(row) else None
        if numero_raw is None or str(numero_raw).strip() == "":
            continue

        numero = _int(numero_raw)
        disciplina = str(row[col["disciplina"]]).strip() if row[col["disciplina"]] is not None else ""
        gabarito = _alt(row[col["gabarito"]])
        acerto = row[col["acerto"]]
        counts = {k: _int(row[col[k]]) for k in ("A", "B", "C", "D", "E")}
        n_resp = sum(counts.values())

        # validações por linha
        if not disciplina:
            erros.append(f"Linha {ridx}: disciplina vazia.")
        if gabarito is None:
            erros.append(f"Linha {ridx}: gabarito inválido ({row[col['gabarito']]!r}); linha ignorada.")
            continue
        try:
            acerto_f = float(acerto)
        except (TypeError, ValueError):
            erros.append(f"Linha {ridx}: 'Acertos Escola' inválido ({acerto!r}); linha ignorada.")
            continue
        if not (0 <= acerto_f <= 1):
            erros.append(f"Linha {ridx}: 'Acertos Escola'={acerto_f} fora de [0,1]; linha ignorada.")
            continue
        if n_resp == 0:
            erros.append(f"Linha {ridx}: soma A–E = 0 (nenhum respondente).")

        questoes.append({
            "numero": numero,
            "disciplina": disciplina,
            "gabarito": gabarito,
            "acerto_pct": round(acerto_f * 100, 2),
            "n_respondentes": n_resp,
            "A": counts["A"], "B": counts["B"], "C": counts["C"],
            "D": counts["D"], "E": counts["E"],
            "item_1": _alt(row[col["item_1"]]) if "item_1" in col and col["item_1"] < len(row) else None,
            "item_2": _alt(row[col["item_2"]]) if "item_2" in col and col["item_2"] < len(row) else None,
            "item_3": _alt(row[col["item_3"]]) if "item_3" in col and col["item_3"] < len(row) else None,
        })

    if not questoes:
        raise ValueError("Nenhuma questão válida encontrada na planilha.")

    return questoes, erros
